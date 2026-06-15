import streamlit as st
import pandas as pd
import numpy as np
import io
import os
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

# ==========================================
# 輔助函式：手機號碼格式化
# ==========================================
def normalize_phone(val):
    """
    將手機號碼轉為字串，去除 .0，並確保 09 開頭
    """
    if pd.isna(val) or val == "":
        return ""
    
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if len(s) == 9 and s.startswith("9"):
        s = "0" + s
        
    return s

# ==========================================
# 頁面基本設定
# ==========================================
st.set_page_config(page_title="自動對帳系統 (最終升級版)", page_icon="📊", layout="wide")
st.title("📊 自動對帳系統")

mode = st.sidebar.radio("請選擇對帳功能：", ["🚗 洗車與三合一對帳 (Code A)", "📺 LiTV 對帳 (Code B)"])

# ==========================================
# 🚗 功能 A：洗車與三合一對帳邏輯 (分離上傳、限定欄位、新制車牌)
# ==========================================
def process_car_wash(files_wash_a, files_3in1_a, file_billing_upload, match_mode):
    output = io.BytesIO()
    logs = []
    output_filename = "洗車與三合一_對帳結果.xlsx"

    try:
        if file_billing_upload:
            base_name = os.path.splitext(file_billing_upload.name)[0]
            output_filename = f"{base_name}_CMX確認.xlsx"

        file_billing_upload.seek(0)

        col_id = '訂單編號'
        col_plate = '車牌'
        col_refund = '退款時間'
        col_phone = '手機號碼'
        target_month_str = datetime.now().strftime("%Y/%m")

        # ---------------------------------------------------------
        # 1. 定義 A 表讀取與清理邏輯
        # ---------------------------------------------------------
        def prepare_a_data(file_list, label):
            if not file_list: 
                return pd.DataFrame(), pd.DataFrame()
                
            logs.append(f"📂 正在讀取【{label} A表】，共 {len(file_list)} 份檔案...")
            df_list = []
            for f in file_list:
                f.seek(0)
                df_temp = pd.read_excel(f, sheet_name=0, header=2)
                df_list.append(df_temp)
                logs.append(f"   ↳ 成功讀取: {f.name} ({len(df_temp)} 筆)")
                
            df_raw = pd.concat(df_list, ignore_index=True)

            df_ref = pd.DataFrame()
            if col_refund in df_raw.columns:
                df_ref = df_raw[df_raw[col_refund].notna()].copy()
                df_filtered = df_raw[df_raw[col_refund].isna()]
            else:
                df_filtered = df_raw

            df_cln = df_filtered.dropna(subset=[col_id]).copy()
            df_cln[col_id] = df_cln[col_id].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

            if col_plate in df_cln.columns:
                df_cln[col_plate] = df_cln[col_plate].astype(str).str.replace(r'[-\s]', '', regex=True).str.upper()
            else:
                df_cln[col_plate] = ""

            if col_phone not in df_cln.columns:
                df_cln[col_phone] = ""
            else:
                df_cln[col_phone] = df_cln[col_phone].apply(normalize_phone)

            # 依據模式產生「比對用車牌」
            if match_mode == "廠商新制 (手機後7碼-車牌)":
                df_cln['比對用車牌'] = df_cln.apply(
                    lambda r: f"{str(r[col_phone])[-7:]}-{r[col_plate]}" if len(str(r[col_phone])) >= 7 else f"{r[col_phone]}-{r[col_plate]}", 
                    axis=1
                )
            else:
                df_cln['比對用車牌'] = df_cln[col_plate]

            df_cln = df_cln.drop_duplicates(subset=[col_id, '比對用車牌'])
            logs.append(f"   ↳ 【{label} A表】合併去重後，共 {len(df_cln)} 筆有效資料")
            return df_cln, df_ref

        # 讀取洗車與三合一的 A 表
        df_a_wash, df_ref_wash = prepare_a_data(files_wash_a, "洗車")
        df_a_3in1, df_ref_3in1 = prepare_a_data(files_3in1_a, "三合一")
        
        # 合併兩邊的退款名單
        df_a_refunds = pd.concat([df_ref_wash, df_ref_3in1], ignore_index=True)
        if match_mode == "廠商新制 (手機後7碼-車牌)":
             logs.append("   ⚠️ 已啟用新制：A表比對鍵轉換為【手機後7碼-車牌】格式")

        # ---------------------------------------------------------
        # 2. 處理右側檔案 (請款明細 / B表) - 動態抓取工作表
        # ---------------------------------------------------------
        logs.append(f"📂 正在讀取右側檔案 (請款明細/B表)...")
        xls_b = pd.ExcelFile(file_billing_upload)
        available_sheets = xls_b.sheet_names
        
        sheet_name_billing = '請款' if '請款' in available_sheets else available_sheets[0]

        wash_candidates = [s for s in available_sheets if ('明細' in s or 'detail' in s.lower()) and '三合一' not in s]
        sheet_name_wash = wash_candidates[0] if wash_candidates else (available_sheets[1] if len(available_sheets)>1 else available_sheets[0])
        
        three_in_one_candidates = [s for s in available_sheets if '三合一' in s]
        sheet_name_3in1 = three_in_one_candidates[0] if three_in_one_candidates else None
        
        logs.append(f"   🚀 B表鎖定工作表 ➔ 摘要: '{sheet_name_billing}' | 洗車: '{sheet_name_wash}' | 三合一: '{sheet_name_3in1}'")

        # --- 讀取摘要表 ---
        df_temp = pd.read_excel(xls_b, sheet_name=sheet_name_billing, header=None, usecols="A:E", nrows=20)
        header_row_idx = 2
        for i, row in df_temp.iterrows():
            if '提供日期' in " ".join([str(x) for x in row.values]):
                header_row_idx = i
                break
        
        df_daily = pd.read_excel(xls_b, sheet_name=sheet_name_billing, header=header_row_idx, usecols="A:E")
        if len(df_daily.columns) >= 5:
            val_count = pd.to_numeric(df_daily.iloc[:, 1], errors='coerce').fillna(0).sum()
            val_billing = pd.to_numeric(df_daily.iloc[:, 2], errors='coerce').fillna(0).sum()
            val_sms = pd.to_numeric(df_daily.iloc[:, 4], errors='coerce').fillna(0).sum()
            val_total = val_billing + val_sms
        else:
            val_count, val_billing, val_sms, val_total = 0, 0, 0, 0

        # ---------------------------------------------------------
        # 3. 定義子對帳函式 (限定必要欄位！)
        # ---------------------------------------------------------
        def merge_datasets(df_a_sub, sheet_name_b):
            if df_a_sub.empty or not sheet_name_b: 
                return pd.DataFrame(), pd.DataFrame()
            
            df_b_raw = pd.read_excel(xls_b, sheet_name=sheet_name_b)
            df_b_sub = df_b_raw.dropna(subset=[col_id]).copy()
            df_b_sub[col_id] = df_b_sub[col_id].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
            df_b_sub = df_b_sub[~df_b_sub[col_id].str.contains('合計|Total|總計', case=False, na=False)]
            
            if col_plate in df_b_sub.columns:
                if match_mode == "廠商新制 (手機後7碼-車牌)":
                    df_b_sub['比對用車牌'] = df_b_sub[col_plate].astype(str).str.upper().str.strip()
                else:
                    df_b_sub['比對用車牌'] = df_b_sub[col_plate].astype(str).str.replace(r'[-\s]', '', regex=True).str.upper()
            else:
                df_b_sub['比對用車牌'] = ""
                
            if col_phone not in df_b_sub.columns:
                df_b_sub[col_phone] = ""
            else:
                df_b_sub[col_phone] = df_b_sub[col_phone].apply(normalize_phone)
                
            df_b_sub = df_b_sub.drop_duplicates(subset=[col_id, '比對用車牌'])
            
            # 🎯 核心修正：只保留必要的欄位去對帳，讓畫面乾淨
            base_cols_keep = [col_id, col_plate, col_phone]
            cols_a = [c for c in base_cols_keep if c in df_a_sub.columns] + ['比對用車牌']
            cols_b = [c for c in base_cols_keep if c in df_b_sub.columns] + ['比對用車牌']
            
            # 去除重複的欄位名
            cols_a = list(dict.fromkeys(cols_a))
            cols_b = list(dict.fromkeys(cols_b))
            
            df_total = pd.merge(
                df_a_sub[cols_a], 
                df_b_sub[cols_b], 
                on=[col_id, '比對用車牌'], 
                how='outer', 
                indicator=True, 
                suffixes=('_A', '_B')
            )
            
            # 丟掉內部的「比對用車牌」欄位，不顯示在報表上
            df_total = df_total.drop(columns=['比對用車牌'], errors='ignore')
            return df_total, df_b_sub

        # 執行兩路對帳
        df_total_wash, df_b_wash_clean = merge_datasets(df_a_wash, sheet_name_wash)
        df_total_3in1, df_b_3in1_clean = merge_datasets(df_a_3in1, sheet_name_3in1)

        logs.append(f"   ↳ 📊 B表有效筆數統計：洗金寶 {len(df_b_wash_clean)} 筆，三合一 {len(df_b_3in1_clean)} 筆")
        logs.append(f"✅ 雙路對帳完成！輸出報表已限定必填欄位。")

        # ---------------------------------------------------------
        # 4. 寫入 Excel (模組化寫入)
        # ---------------------------------------------------------
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            wb = writer.book
            
            fmt_header = wb.add_format({'bold': True, 'bg_color': '#EFEFEF', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 12})
            fmt_content = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 11})
            fmt_currency = wb.add_format({'num_format': '#,##0', 'border': 1, 'align': 'right', 'valign': 'vcenter', 'font_size': 11})
            fmt_blue = wb.add_format({'bg_color': '#DDEBF7', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 11})
            fmt_pink = wb.add_format({'bg_color': '#FCE4D6', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 11})

            # (1) 寫入摘要表
            ws1 = wb.add_worksheet('請款')
            top_headers = ['統計月份', '轉檔筆數', '轉檔請款金額', '簡訊請款金額', '合計金額']
            top_values = [target_month_str, val_count, val_billing, val_sms, val_total]
            ws1.set_row(0, 30)
            ws1.set_row(1, 25)
            for col, (header, val) in enumerate(zip(top_headers, top_values)):
                ws1.write(0, col, header, fmt_header)
                if col == 0: ws1.write(1, col, val, fmt_content)
                else: ws1.write(1, col, val, fmt_currency)
            for col_idx, col_name in enumerate(df_daily.columns):
                ws1.write(3, col_idx, col_name, fmt_header)
            for r, row in enumerate(df_daily.values):
                for c, val in enumerate(row):
                    if pd.isna(val): write_val = ""
                    else: write_val = val
                    ws1.write(r + 4, c, write_val, fmt_content)
            ws1.set_column('A:E', 25) 

            # (2) 定義副程式：負責寫入對帳表群組
            def write_result_sheets(df_result, prefix_name):
                if df_result.empty: return
                ws = wb.add_worksheet(f'{prefix_name}_對帳總表')
                columns = df_result.columns.tolist()
                for c_idx, col_name in enumerate(columns):
                    ws.write(0, c_idx, col_name, fmt_header)
                ws.set_column(0, len(columns)-1, 22)
                ws.set_row(0, 22)

                for r_idx, row in df_result.iterrows():
                    merge_status = row['_merge']
                    if merge_status == 'left_only': current_fmt = fmt_blue
                    elif merge_status == 'right_only': current_fmt = fmt_pink
                    else: current_fmt = fmt_content
                    
                    excel_row = r_idx + 1
                    ws.set_row(excel_row, 18) 
                    for c_idx, val in enumerate(row):
                        write_val = "" if pd.isna(val) else val
                        ws.write(excel_row, c_idx, write_val, current_fmt)

                df_result[df_result['_merge'] == 'left_only'].drop(columns=['_merge']).to_excel(writer, sheet_name=f'{prefix_name}_僅A表有', index=False)
                df_result[df_result['_merge'] == 'right_only'].drop(columns=['_merge']).to_excel(writer, sheet_name=f'{prefix_name}_僅B表有', index=False)

            # (3) 寫入洗車結果
            write_result_sheets(df_total_wash, "洗車")
            
            # (4) 寫入三合一結果
            write_result_sheets(df_total_3in1, "三合一")
            
            # (5) 寫入退款排除
            if not df_a_refunds.empty:
                df_a_refunds.to_excel(writer, sheet_name='A表退款排除名單', index=False)

        return output.getvalue(), logs, output_filename

    except Exception as e:
        import traceback
        return None, [f"❌ 錯誤: {str(e)}", traceback.format_exc()], None

# ==========================================
# 📺 功能 B：LiTV 對帳邏輯 (維持不變)
# ==========================================
def process_litv(file_a_upload, file_b_upload):
    output_buffer = io.BytesIO()
    logs = []
    output_filename = "LiTV_CMX確認.xlsx"

    try:
        xl_a = pd.ExcelFile(file_a_upload)
        xl_b = pd.ExcelFile(file_b_upload)
        file_a_target, file_b_target = file_a_upload, file_b_upload

        if 'ACG對帳明細' in xl_a.sheet_names and 'ACG對帳明細' not in xl_b.sheet_names:
            logs.append("💡 偵測到檔案順序相反，已自動交換 A/B 表。")
            file_a_target, file_b_target = file_b_upload, file_a_upload
        elif 'ACG對帳明細' in xl_b.sheet_names:
            logs.append("✅ 檔案順序正確。")
        else:
             return None, [f"❌ 錯誤：找不到「ACG對帳明細」。"], None, None, None
        
        base_name = os.path.splitext(file_b_target.name)[0]
        output_filename = f"{base_name}_CMX確認.xlsx"
        
        file_a_target.seek(0)
        file_b_target.seek(0)

        logs.append("正在載入 B 表...")
        wb = openpyxl.load_workbook(file_b_target)

        logs.append("正在讀取 A 表 (header=2)...")
        df_a = pd.read_excel(file_a_target, header=2)
        df_a.columns = df_a.columns.str.strip()
        
        if '金額' not in df_a.columns: return None, [f"❌ 錯誤：A 表讀不到「金額」欄位。"], None, None, None

        df_a['金額'] = pd.to_numeric(df_a['金額'], errors='coerce').fillna(0)
        df_a_filtered = df_a[(df_a['金額'] > 0) & (df_a['退款時間'].isna()) & (df_a['手機號碼'].notna())].copy()

        def fix_phone_a(val):
            if pd.isna(val): return ""
            s = str(val).split('.')[0]
            if len(s) == 9: s = '0' + s
            return s

        df_a_filtered['手機全碼'] = df_a_filtered['手機號碼'].apply(fix_phone_a)
        df_a_filtered['手機隱碼'] = df_a_filtered['手機全碼'].apply(lambda x: x[:6] + '****' if len(x) >= 10 else x)
        a_lookup_set = set(zip(df_a_filtered['手機隱碼'], df_a_filtered['方案(SKU)'].str.strip()))

        logs.append("正在讀取 ACG 對帳明細...")
        file_b_target.seek(0)
        df_b_acg_full = pd.read_excel(file_b_target, sheet_name='ACG對帳明細')
        df_b_acg_full.columns = df_b_acg_full.columns.str.strip()

        stop_idx = next((idx for idx, val in enumerate(df_b_acg_full['編號']) if "不計費" in str(val)), None)
        df_b_valid = df_b_acg_full.iloc[:stop_idx].copy() if stop_idx is not None else df_b_acg_full.copy()
        
        df_b_valid = df_b_valid.dropna(subset=['手機/虛擬帳號', '廠商對帳key1']).copy()
        df_b_valid['手機/虛擬帳號'] = df_b_valid['手機/虛擬帳號'].astype(str).str.strip()
        df_b_valid['廠商對帳key1'] = df_b_valid['廠商對帳key1'].astype(str).str.strip()
        b_lookup_set = set(zip(df_b_valid['手機/虛擬帳號'], df_b_valid['廠商對帳key1']))

        sku_mapping = {'LiTV_LUX_1Y_OT': ['LiTV_LUX_1Y_OT', 'LiTV_LUX_F1MF_1Y_OT'], 'LiTV_LUX_1M_OT': ['LiTV_LUX_1M_OT']}
        reverse_sku_map = {'LiTV_LUX_F1MF_1Y_OT': 'LiTV_LUX_1Y_OT', 'LiTV_LUX_1Y_OT': 'LiTV_LUX_1Y_OT', 'LiTV_LUX_1M_OT': 'LiTV_LUX_1M_OT'}

        sheet1_data, diff_a_not_b = [], []
        for _, row in df_a_filtered.iterrows():
            sku_a, phone_masked = str(row['方案(SKU)']).strip(), row['手機隱碼']
            found_in_b = any((phone_masked, k) in b_lookup_set for k in sku_mapping.get(sku_a, [sku_a]))

            if sku_a == 'LiTV_LUX_1M_OT': out_sku, out_amt, out_name = 'LiTV_LUX_1M_OT', 187, '豪華雙享餐/月繳/單次(定價$250)'
            elif sku_a == 'LiTV_LUX_1Y_OT': out_sku, out_amt, out_name = 'LiTV_LUX_F1MF_1Y_OT', 1717, '豪華雙享餐-首月免費/年繳/單次(定價$2,290)'
            else: out_sku, out_amt, out_name = sku_a, row['金額'], sku_a

            sheet1_data.append({'廠商方案代碼': out_sku, '廠商方案名稱': out_name, '手機/虛擬帳號': phone_masked, '方案金額': out_amt, 'CMX訂單編號': row['訂單編號'], 'is_diff': not found_in_b})
            if not found_in_b: diff_a_not_b.append({'手機號碼': row['手機全碼'], '方案': sku_a, '訂單編號': row['訂單編號']})

        diff_b_not_a = []
        for _, row in df_b_valid.iterrows():
            b_phone, b_key = str(row['手機/虛擬帳號']).strip(), str(row['廠商對帳key1']).strip()
            if "*" in b_phone and (b_phone, reverse_sku_map.get(b_key, b_key)) not in a_lookup_set:
                diff_b_not_a.append({'手機/虛擬帳號': b_phone, '廠商對帳key1': b_key})

        logs.append("正在寫入 Excel...")
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        font_style = Font(size=18)

        if "CMX對帳明細" in wb.sheetnames: del wb["CMX對帳明細"]
        ws_new = wb.create_sheet("CMX對帳明細", 0)
        headers = ['廠商方案代碼', '廠商方案名稱', '手機/虛擬帳號', '方案金額', 'CMX訂單編號']
        ws_new.append(headers)
        
        for data in sheet1_data:
            ws_new.append([data[h] for h in headers])
            for cell in ws_new[ws_new.max_row]:
                cell.font = font_style
                if data['is_diff']: cell.fill = yellow_fill

        if 'ACG對帳明細' in wb.sheetnames:
            ws_acg = wb['ACG對帳明細']
            h_list = [cell.value for cell in ws_acg[1]]
            if '手機/虛擬帳號' in h_list and '廠商對帳key1' in h_list:
                p_idx, k_idx = h_list.index('手機/虛擬帳號') + 1, h_list.index('廠商對帳key1') + 1
                max_row = (stop_idx + 1) if stop_idx is not None else ws_acg.max_row
                for r_idx in range(2, max_row + 1):
                    p_val, k_val = str(ws_acg.cell(row=r_idx, column=p_idx).value).strip(), str(ws_acg.cell(row=r_idx, column=k_idx).value).strip()
                    for cell in ws_acg[r_idx]: cell.font = font_style
                    if "*" in p_val and (p_val, reverse_sku_map.get(k_val, k_val)) not in a_lookup_set:
                        for cell in ws_acg[r_idx]: cell.fill = yellow_fill
        
        wb.save(output_buffer)
        return output_buffer.getvalue(), logs, diff_a_not_b, diff_b_not_a, output_filename

    except Exception as e:
        return None, [f"❌ 程式執行錯誤: {str(e)}"], None, None, None


# ==========================================
# 介面顯示邏輯
# ==========================================

if mode == "🚗 洗車與三合一對帳 (Code A)":
    st.header("🚗 洗車與三合一 聯合對帳")
    
    match_mode = st.radio(
        "⚙️ 請選擇廠商車牌比對模式：", 
        ["預設模式 (純車牌比對)", "廠商新制 (手機後7碼-車牌)"],
        horizontal=True
    )
    
    st.info("💡 邏輯：將【洗車】與【三合一】的 A 表分開上傳，系統會自動核對同一張 B 表裡的不同明細。")
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("<h3 style='text-align: center; color: #E74C3C;'>1. CMX報表 (A表上傳區)</h3>", unsafe_allow_html=True)
        
        st.markdown("**🚗 洗車 A 表 (支援多選)**")
        files_wash = st.file_uploader(" ", type=['xlsx', 'xls'], key="wash_supplier", label_visibility="collapsed", accept_multiple_files=True)
        
        st.markdown("**📦 三合一 A 表 (支援多選，若無則免傳)**")
        files_3in1 = st.file_uploader(" ", type=['xlsx', 'xls'], key="3in1_supplier", label_visibility="collapsed", accept_multiple_files=True)
    
    with col2:
        st.markdown("<h3 style='text-align: center; color: #2E86C1;'>2. TMS請款明細 (B表)</h3>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; color: transparent;'>僅限單一檔案</p>", unsafe_allow_html=True)
        file_billing = st.file_uploader(" ", type=['xlsx', 'xls'], key="car_billing", label_visibility="collapsed")
    
    if st.button("🚀 開始自動對帳", type="primary"):
        # 只要有一種 A表 有上傳，並且 B表 有上傳，即可啟動
        if (len(files_wash) > 0 or len(files_3in1) > 0) and file_billing:
            with st.spinner("資料比對中，請稍候..."):
                result, logs, filename = process_car_wash(files_wash, files_3in1, file_billing, match_mode)
            
            st.expander("執行紀錄 (點擊展開)", expanded=True).write(logs)
            
            if result:
                st.success("🎉 對帳完成！")
                st.download_button(
                    label=f"📥 下載結果 ({filename})",
                    data=result,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("⚠️ 請確認「至少一份 A 表」與「B 表」都已完成上傳。")

elif mode == "📺 LiTV 對帳 (Code B)":
    st.header("📺 LiTV 訂單對帳")
    st.info("💡 邏輯：A表讀 header=2，B表找 ACG對帳明細")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("<h3 style='text-align: center; color: #E74C3C;'>1. CMX報表 (A表)</h3>", unsafe_allow_html=True)
        file_a = st.file_uploader(" ", type=['xlsx', 'xls'], key="litv_a", label_visibility="collapsed")
    with col2:
        st.markdown("<h3 style='text-align: center; color: #2E86C1;'>2.  LiTV請款明細  (B表)</h3>", unsafe_allow_html=True)
        file_b = st.file_uploader(" ", type=['xlsx', 'xls'], key="litv_b", label_visibility="collapsed")
    
    if st.button("🚀 開始 LiTV 對帳", type="primary"):
        if file_a and file_b:
            with st.spinner("LiTV 資料比對中..."):
                result, logs, diff_a, diff_b, filename = process_litv(file_a, file_b)
            
            with st.expander("執行紀錄", expanded=True):
                for l in logs: st.text(l)
            
            if result:
                st.success("成功！")
                c1, c2 = st.columns(2)
                c1.error(f"A有B無 (共 {len(diff_a) if diff_a else 0} 筆)")
                if diff_a: c1.dataframe(pd.DataFrame(diff_a))
                c2.warning(f"B有A無 (共 {len(diff_b) if diff_b else 0} 筆)")
                if diff_b: c2.dataframe(pd.DataFrame(diff_b))
                st.download_button(label=f"📥 下載結果 ({filename})", data=result, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning("⚠️ 請確認兩個檔案都已上傳。")
